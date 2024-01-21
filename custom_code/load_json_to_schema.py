import pandas as pd

import json

from json import JSONDecodeError
from genson import SchemaBuilder

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill


class Make_Schema:
	def __int__(self, json_file):
		self.json_file = json_file
	# 엑셀 색칠 및 선 그리기
	def fill_excel(self, excel_file):
		wb = load_workbook(excel_file)
		ws = wb['샘플진단규칙예시']

		# D2:I24까지 노란색으로 칠하기
		yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
		for row in range(2, 25):
			for col in range(4, 10):
				cell = ws.cell(row=row, column=col)
				cell.fill = yellow_fill

		# A1:I24까지 표 선 그리기
		thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
		                     bottom=Side(style='thin'))
		for row in range(1, 25):
			for col in range(1, 10):
				cell = ws.cell(row=row, column=col)
				cell.border = thin_border

		# 워크북 저장
		wb.save(excel_file)
	# 스키마 구조 생성
	def schema_builder(self):
		builder = SchemaBuilder()
		with open(self.json_file, 'r', encoding='utf-8') as f:
			try:
				datastore = json.load(f)
				builder.add_object(datastore)
			except JSONDecodeError as e:
				assert False, 'Json형식이 맞지 않습니다.'
		schema = builder.to_schema()
		schema.pop('$schema')

		assert schema['type'] != 'array', f'Array로 시작하는 파일은 권장하지 않는 형식입니다. 해당 형식으로 원하실 경우 문의해주시길 바랍니다.'
		assert schema['type'] != 'string', f'String 시작하는 파일은 권장하지 않는 형식입니다. 해당 형식으로 원하실 경우 문의해주시길 바랍니다.'
		assert schema['type'] != 'integer', f'Integer 시작하는 파일은 권장하지 않는 형식입니다. 해당 형식으로 원하실 경우 문의해주시길 바랍니다.'
		assert schema['type'] != 'number', f'Number 시작하는 파일은 권장하지 않는 형식입니다. 해당 형식으로 원하실 경우 문의해주시길 바랍니다.'

		df = self.get_normalized_df(schema)
		return df

	# 스키마 바탕으로 DF생성
	def get_normalized_df(self, schema):
		fin_df = pd.json_normalize(schema).transpose()[2:].reset_index()
		fin_df.columns = ['항목', '타입']
		return fin_df
	# DF 타입 변경 및 검출
	def is_array_object(self, idx, temp_list, fin_df, temp_array_name, new_item: list, new_type: list, del_list: list):
		if len(temp_list) > 2:
			if fin_df['타입'].iloc[idx] == 'array':
				temp_array_name = temp_list[-1]

			if temp_array_name != temp_list[-1] and temp_array_name is not None \
					and fin_df['타입'].iloc[idx] == 'object' and temp_list[-1] == 'items':
				new_name = f"{temp_list[0]}.{temp_array_name}"
				if new_name in new_item:
					num_idx = new_item.index(new_name)
					if new_type[num_idx] == 'array':
						new_type[num_idx] = 'array/object'
						del_list.append(idx)
						return True

			elif temp_list[-1] == 'items':
				new_name = f'{temp_list[0]}.{temp_list[-2]}'
				if new_name in new_item:
					num_idx = new_item.index(new_name)
					if new_type[num_idx] == 'array':
						new_type[num_idx] = 'array/object'
						del_list.append(idx)
						return True
				else:
					print('수정해라')

			elif (temp_list[-2] == 'properties' and temp_list[-3] == 'items') or (
					temp_list[1] == 'items' and temp_list[2] == 'properties') or (temp_list[1] == 'properties'):
				temp_list = [x for x in temp_list if x not in ['properties', 'items']]
				new_name = '.'.join(temp_list)
				new_item.append(new_name)
				new_type.append(fin_df['타입'].iloc[idx])
				return True
		return False
	# 최종 DF 생성 및 Excel 생성
	def process_df(self):
		df = self.schema_builder()

		del_list = []
		new_item = []
		new_type = []
		temp_array_name = None

		df['필수여부'] = 'Y'
		df['유효값'] = None
		df['최솟값'] = None
		df['최댓값'] = None
		df['패턴'] = None
		df['기타'] = None
		df['Null 여부'] = None

		for idx, item in enumerate(df['항목']):
			assert item.split('.')[-1] != 'anyOf', f'{item.split(".")[1]}부분은 권장하지 않는 형식입니다. 해당 형식으로 원하실 경우 문의해주시길 바랍니다.'
			if item.split('.')[0] == 'properties' and item.split('.')[-1] == 'required':
				del_list.append(idx)
			elif item.split('.')[0] == 'properties' and item.split('.')[-1] == 'type':
				if '.'.join(item.split('.')[1:-1]).split('.')[-1] == 'items' and df['타입'].iloc[idx] not in ['array',
				                                                                                            'object']:
					del_list.append(idx)
				else:
					assert type(df.iloc[idx][
						            '타입']) == str, f"{item.split('.')[-2]} 타입이 {len(df.iloc[idx]['타입'])}개 존재하여 Json 형식이 맞지 않습니다."
					temp_list = item.split('.')[1:-1]
					if not self.is_array_object(idx, temp_list, df, temp_array_name, new_item, new_type, del_list):
						new_name = item.split('.')[1]
						if new_name in new_item:
							num_idx = new_item.index(new_name)

							if new_type[num_idx] == 'array':
								new_type[num_idx] = 'array/object'
								del_list.append(idx)
						else:
							new_item.append(new_name)
							new_type.append(df['타입'].iloc[idx])

		df.drop(del_list, axis=0, inplace=True)
		df['항목'] = new_item
		df['타입'] = new_type

		excel_file = 'validate.xlsx'
		writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')

		sample = pd.read_excel('sample.xlsx')
		df.to_excel(writer, index=False, header=True, sheet_name='샘플진단규칙')
		sample.to_excel(writer, index=False, header=True, sheet_name='샘플진단규칙예시')

		writer.save()
		self.fill_excel(excel_file)

		return df


if __name__ =='__main__':
	schema = Make_Schema()
	schema.json_file = '/data/Model/uni_solution/object_detection/test.json'
	df = schema.process_df()
	# print(df)
