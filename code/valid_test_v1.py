from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection

wb = load_workbook('last_test.xlsx')
worksheet = wb['샘플진단규칙예시']

# D2:I24까지 노란색으로 칠하기
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for row in range(2, 25):
    for col in range(4, 10):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = yellow_fill

# A1:I24까지 표 선 그리기
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row in range(1, 25):
    for col in range(1, 10):
        cell = worksheet.cell(row=row, column=col)
        cell.border = thin_border

# 워크북 저장
wb.save('last_test.xlsx')






