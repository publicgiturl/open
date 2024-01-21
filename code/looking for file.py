import glob, os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import json
from openpyxl import Workbook
from typing_extensions import final
main_path = 'Z:/학습데이터/[066]20.11.03 하남감일(경기도 하남시 감이동)_D02_66'


os.chdir(main_path)
before_path = main_path
after_path = main_path
dir_path = main_path.replace(main_path.split('/')[-1], '')

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)
 
createFolder(dir_path+'카운트')


def start():
    write_xl = Workbook()
    write_ws = write_xl.active
    write_ws.append(['폴더명', 'JSON', 'box', 'polygon'])
    for (path, dir, files) in os.walk(before_path):
        firstRow = True
        for fileName in files:
            if fileName.endswith('.json'):
                firstJson = True
                try:
                    with open(os.path.join(path, fileName), encoding='utf-8') as json_file:
                        jsonData = json.load(json_file)
                        deleteList = []
                        for i in range(len(jsonData['annotations'])):
                            if 'box' in jsonData['annotations'][i]:
                                if jsonData['annotations'][i]['class'] == '03':
                                    deleteList.append(i)
                                    
                            if 'box' in jsonData['annotations'][i]:
                                xNum = jsonData['annotations'][i]['box'][2] - jsonData['annotations'][i]['box'][0]
                                yNum = jsonData['annotations'][i]['box'][3] - jsonData['annotations'][i]['box'][1]

                                if xNum <= 40 and yNum <= 40:
                                    deleteList.append(i)
                                    if firstRow and firstJson:
                                        write_ws.append([path, fileName, str(jsonData['annotations'][i]['box']).replace('[', '').replace(']', '')])
                                        firstRow = False
                                        firstJson = False
                                    elif firstJson:
                                        write_ws.append(['', fileName, str(jsonData['annotations'][i]['box']).replace('[', '').replace(']', '')])
                                        firstJson = False
                                    else:
                                        write_ws.append(['', '', str(jsonData['annotations'][i]['box']).replace('[', '').replace(']', '')])
                                
                            elif 'polygon' in jsonData['annotations'][i]:
                                if len(jsonData['annotations'][i]['polygon']) < 5 or not jsonData['annotations'][i]['class'] in ['03', '04']:

                                    deleteList.append(i)

                                    write_ws.append([path, fileName,'', str(jsonData['annotations'][i]['polygon'])])
        
                                
                        for k in range(len(deleteList)):
                            jsonData['annotations'].pop(deleteList[k] - k)


                        with open(os.path.join(after_path, fileName), 'w', encoding='utf-8') as json_change:
                            json.dump(jsonData, json_change, indent=2, ensure_ascii=False)
                except Exception as ex:
                    print(ex)

                finally:
                    json_file.close()

    write_xl.save(os.path.join(main_path, 'delete_data.xlsx'))

if __name__ == '__main__':
    start()

def Start():
    for (path, dir, files) in os.walk(before_path):
        for file_name in files:
            if file_name.endswith('.json'):
                try:
                    with open(os.path.join(path, file_name), encoding='utf-8') as json_file:
                        json_data = json.load(json_file)
                        annotation = json_data['annotations']
                        error_annotation_list = []
                        

                        for i in range(len(annotation)):
                            if 'box' in annotation[i]:
                                pass
                            elif 'polygon' in annotation[i]:
                                pass
                            elif 'keypoints' in annotation[i]:
                                pass
                            else:
                                error_annotation_list.append(i)
 
                        for i in range(len(error_annotation_list)):
                            json_data['annotations'].pop(error_annotation_list[i] - i)

                        with open(os.path.join(after_path, file_name), 'w', encoding='utf-8') as json_change:
                            json.dump(json_data, json_change, indent=2)

                finally:
                    json_file.close()


if __name__ == '__main__':
    Start()

def start():
    for (path, dir, files) in os.walk(main_path):
        for file_name in files:
            if file_name.endswith('.json'):
                try:
                    with open(os.path.join(path, file_name), encoding='utf-8') as json_file:
                        jsonData = json.load(json_file)
                        jsonData['image']['copyrighter'] = '미디어그룹사람과숲(컨)'
                        for arr in range(len(jsonData['annotations'])):
                            jsonData['annotations'][arr]['flags'] = {} 
                            
                    with open(os.path.join(path, file_name), 'w', encoding='utf-8') as json_change:
                        json.dump(jsonData, json_change, ensure_ascii = False, indent = 4)
                        
                finally:
                    json_file.close()
                    
if __name__ == '__main__':
    start()


def Start():
    for (path, dir, files) in os.walk(main_path):
        error_json = {}
        for file_name in files:
            if file_name.endswith('.json'):
                try:
                    with open(os.path.join(path, file_name), encoding='utf-8') as json_file:
                        error_json[os.path.join(path, file_name)] = ''
                        json_data = json.load(json_file)
                        annotation = json_data['annotations']
                        com_count = 0

                        if json_data['image']['copyrighter'] != '미디어그룹사람과숲(컨)':
                            error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'Copyrighter 에러'
                            com_count += 1
                        if len(json_data['image']['date']) != 8:
                            if com_count != 0:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'date 에러'
                            else:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'date 에러'
                                com_count += 1
                        if json_data['image']['path'] == 'NULL':
                            if com_count != 0:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'path 에러'
                            else:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'path 에러'
                                com_count += 1
                        if json_data['image']['filename'][0] != 'S' or json_data['image']['filename'][3] != 'O' and json_data['image']['filename'][3] != 'N':
                            if com_count != 0:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'filename 에러'
                            else:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'filename 에러'
                                com_count += 1
                        if len(annotation) == 0:
                            if com_count != 0:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'annotation 누락'
                            else:
                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'annotation 누락'
                                com_count += 1
                        else:
                            for i in range(len(annotation)):
                                if 'data ID' not in annotation[i] or 'middle classification' not in annotation[i] or 'class' not in annotation[i]:
                                    if com_count != 0:
                                        error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'annotation 에러'
                                        break
                                    else:
                                        error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'annotation 에러'
                                        com_count += 1
                                        break
                                else:
                                    if len(annotation[i]['data ID']) != 2:
                                        if com_count != 0:
                                            error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'annotation 에러'
                                            break
                                        else:
                                            error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'annotation 에러'
                                            com_count += 1
                                            break
                                    if 'box' in annotation[i]:
                                        if len(annotation[i]['box']) != 4:
                                            if com_count != 0:
                                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'annotation 에러'
                                                break
                                            else:
                                                error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'annotation 에러'
                                                com_count += 1
                                                break
                                    elif 'polygon' in annotation[i]:
                                        pass
                                    elif 'keypoints' in annotation[i]:
                                        pass
                                    else:
                                        if com_count != 0:
                                            error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + ', ' + 'annotation 에러'
                                            break
                                        else:
                                            error_json[os.path.join(path, file_name)] = error_json[os.path.join(path, file_name)] + 'annotation 에러'
                                            com_count += 1
                                            break

                except:
                    error_json[os.path.join(path, file_name)] = 'json 형식 에러'
                finally:
                    json_file.close()

        try:
            e = open(os.path.join(path, 'annotation_error.csv'), mode='wt', encoding='utf-8')
            for key, value in error_json.items():
                if value != '':
                    e.write(str(key) + ' : ' + str(value) + '\n')
        finally:
            e.close()


if __name__ == '__main__':
    Start()

remove_list = []

# annotation누락 찾기
with open('annotation_error.csv', 'r', encoding='utf-8') as f:
    Ann_list = [i for i in f]

for ann in Ann_list:
    if ' annotation 누락' in ann:
        remove_list.append(ann.split('\\')[-1].split(' :')[0])

for i in remove_list:
    os.remove('./'+i)



def start():
    write_xl = Workbook()
    write_ws = write_xl.active
    write_ws.append(['폴더명', 'JSON', 'class', 'class_count', 'box', 'polygon', 'point'])
    for (path, dir, files) in os.walk(main_path):
        firstRow = True
        for fileName in files:
            if fileName.endswith('.json'):
                firstJson = True
                with open(os.path.join(path, fileName), encoding='utf-8') as json_file:
                    try:
                        jsonData = json.load(json_file)
                        annotations = jsonData['annotations']
                        classType = []
                        classTypeList = ['box', 'polygon', 'point']
                        classCountDict = {}
                        classTypeCountDict = {}
                        

                        for i in range(len(annotations)):
                            classType.append(annotations[i]['class'])

                        for key in classType:
                            classCountDict[key] = classType.count(key)
                            classTypeCountDict[key] = {'box': 0, 'polygon': 0, 'point': 0}

                        for key in classCountDict:
                            for i in range(len(annotations)):
                                if annotations[i]['class'] == key:
                                    for item in classTypeList:
                                        if item in annotations[i]:
                                            classTypeCountDict[key][item] += 1

                        for key in classCountDict:
                            if firstRow and firstJson:
                                write_ws.append([path, fileName, key, classCountDict[key],
                                                 classTypeCountDict[key][classTypeList[0]],
                                                 classTypeCountDict[key][classTypeList[1]],
                                                 classTypeCountDict[key][classTypeList[2]]])
                                firstRow = False
                                firstJson = False
                            elif firstJson:
                                write_ws.append(["", fileName, key, classCountDict[key],
                                                 classTypeCountDict[key][classTypeList[0]],
                                                 classTypeCountDict[key][classTypeList[1]],
                                                 classTypeCountDict[key][classTypeList[2]]])
                                firstJson = False
                            else:
                                write_ws.append(["", "", key, classCountDict[key],
                                                 classTypeCountDict[key][classTypeList[0]],
                                                 classTypeCountDict[key][classTypeList[1]],
                                                 classTypeCountDict[key][classTypeList[2]]])
                    except Exception as ex:
                        print(ex)
                        pass

    write_xl.save(os.path.join(main_path, '../카운트/{}.xlsx'.format(main_path.split('/')[-1])))


start()

jpg  = []
json = []
NG = []
no_jpg = []
# json파일 이름명 추출
### 주소만 변경 ###

# os.chdir(main_path)
for file in glob.glob("*.json"):
    json.append(file.split('.')[0])

# jpg파일 이름명 추출    
for file in glob.glob("*.jpg"):
    jpg.append(file.split('.')[0])

# Json은 있지만 이미지가 없는 파일 걸러내기
for v in jpg:
    if v not in json:
        NG.append(v+'.jpg')

# 이미지가 없는데 Json이 있는 파일 걸러내기
for v in json:
    if v not in jpg:
        no_jpg.append(v+'.json')

# # 매치 되지않은 jpg리스트 저장
with open('./omission_jpg.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(NG)

# 매치 되지않은 json리스트 저장
with open('./omission_json.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(no_jpg)

# 매치되지 않은 jpg 삭제
for i in NG:
    os.remove(i)

# 매치되지 않은 json 삭제
for i in no_jpg:
    os.remove(i)


print(main_path + '완료')